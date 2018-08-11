import random as rnd
from openpyxl import Workbook
from numpy import longdouble, histogram, std
from numpy.lib import emath


class BaseSet(object):
	def __init__(self, begin, end, seed=0):
		self.begin = begin
		self.end = end
		self.baseset = [longdouble(x) for x in range(begin, end + 1)]
		self.logset = [emath.log10(x) for x in self.baseset]
		rnd.seed(seed)

	def getTotals(self, size):
		total_sum = longdouble(0.0)
		total_product = longdouble(0.0)
		for _ in xrange(size):
			i = rnd.randint(0, len(self.baseset) - 1)
			total_sum += self.baseset[i]
			total_product += self.logset[i]
		return total_sum, total_product


class Results(object):
	def __init__(self, numtrials, begin, end, size):
		self.numtrials = numtrials
		self.begin = begin
		self.end = end
		self.size = size
		self.baseset = BaseSet(self.begin, self.end)
		self.totals = []
		self.histsum = []
		self.binsum = []
		self.histprod = []
		self.binprod = []
		self.stdsum = 0.0
		self.stdprod = 0.0

	def getTotals(self):
		print 'Computing totals...'
		for _ in xrange(self.numtrials):
			self.totals.append((self.baseset.getTotals(self.size)))

	def histogram(self):
		totalsum, totalprod = tuple(zip(*self.totals))
		self.histsum, self.binsum = histogram(totalsum, bins=20)
		self.histprod, self.binprod = histogram(totalprod, bins=20)
		self.stdsum = std(totalsum)
		self.stdprod = std(totalprod)

	def output(self, fname):
		wb = Workbook()
		ws = wb.active

		row = 1
		ws.cell(row=1, column=1, value='total_sum')
		ws.cell(row=1, column=2, value='total_product')
		ws.cell(row=row, column=4, value='hist_sum_bin')
		ws.cell(row=row, column=5, value='hist_sum_values')
		ws.cell(row=row, column=7, value='hist_prod_bin')
		ws.cell(row=row, column=8, value='hist_prod_values')

		row = 2
		for total in self.totals:
			ws.cell(row=row, column=1, value=total[0])
			ws.cell(row=row, column=2, value=total[1])
			row += 1

		row = 2
		for val in range(len(self.histsum)):
			ws.cell(row=row, column=4, value=self.binsum[val])
			ws.cell(row=row, column=5, value=self.histsum[val])
			row += 1

		row = 2
		for val in range(len(self.histprod)):
			ws.cell(row=row, column=7, value=self.binprod[val])
			ws.cell(row=row, column=8, value=self.histprod[val])
			row += 1

		ws.cell(row=1, column=10, value='std_sum')
		ws.cell(row=1, column=11, value=self.stdsum)

		ws.cell(row=2, column=10, value='std_prod')
		ws.cell(row=2, column=11, value=self.stdprod)

		print 'saving {0}'.format(fname)
		wb.save(fname)


if __name__ == '__main__':
	results = Results(100000, 10, 20, 10000)
	results.getTotals()
	results.histogram()
	results.output(r'F:\dev\rationalmodel\src\masa\dist 10 a 20 100000-10000.xlsx')
