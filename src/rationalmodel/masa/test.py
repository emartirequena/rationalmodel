import json
import os
import openpyxl as xl

def main(n: int):
    with open(f'test_1D_N{n}.json', 'rt') as fp:
        content = json.load(fp)

    wb = xl.Workbook()

    # for space in content['spaces']:
    #     ws = wb.create_sheet(f'sheet_{space}')
    #     if 'accumulate' in space:
    #         continue
    #     row1 = int(space) + 2
    #     for cell in content['spaces'][space]:
    #         ispace = int(space) / 2
    #         row = int(cell['pos'][1] + ispace) + 1
    #         column = int(cell['pos'][0] + ispace) + 1
    #         print(ispace, row, column)
    #         ws.cell(row=row, column=column, value=cell['time'])
    #         ws.cell(row=row1 + row, column=column, value=cell['count'])

    for space in content['spaces']:
        if 'accumulate' in space:
            continue
        ws = wb.create_sheet(f'sheet_{space}')
        ispace = int(space) / 2
        for cell in content['spaces'][space]:
            row = int(cell['pos'][0] + ispace) + 1
            ws.cell(row=row, column=1, value=cell['pos'][0])
            ws.cell(row=row, column=2, value=cell['time'])
            count = cell['count']
            ws.cell(row=row, column=3, value=count)
            ws.cell(row=row, column=4, value=cell['next_digits']['0'] / count)
            ws.cell(row=row, column=5, value=cell['next_digits']['1'] / count)

    wb.save(f'test_1D_N{n}.xlsx')

if __name__ == '__main__':
    main(1048575)