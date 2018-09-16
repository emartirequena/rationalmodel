from openpyxl import Workbook


class Results(object):
	def __init__(self, begin, end):
		self.begin = begin
		self.end = end
		self.totals = []

	def getTotals(self):
		print 'Computing totals...'
		for a in range(self.begin, self.end + 1):
			for b in range(a, self.end + 1):
				self.totals.append((a, b, a + b, a * b))

	def output(self, fname):
		wb = Workbook()
		ws = wb.active

		row = 1
		ws.cell(row=row, column=1, value='a')
		ws.cell(row=row, column=2, value='b')
		ws.cell(row=row, column=3, value='sum')
		ws.cell(row=row, column=4, value='product')

		for row in range(len(self.totals)):
			for column in range(4):
				ws.cell(row=row + 2, column=column + 1, value=self.totals[row][column])

		print 'saving {0}'.format(fname)
		wb.save(fname)


if __name__ == '__main__':
	results = Results(2, 100)
	results.getTotals()
	results.output(r'C:\Users\enrique\Google Drive\Enrique\Articulo suma por producto\hojas\dist sum product of two numbers.xlsx')
