import json
import os
from config import Config
from openpyxl import Workbook

from utils import getPeriod


def main():
    config = Config()
    path = config.get('files_path')
    file_paths = [
        os.path.join(path, '1D', 'P06', '1D_N63_P06_F3^2_7.json'),
        os.path.join(path, '1D', 'P08', '1D_N255_P08_F3_5_17.json'),
        os.path.join(path, '1D', 'P10', '1D_N1023_P10_F3_11_31.json'),
        os.path.join(path, '1D', 'P12', '1D_N4095_P12_F3^2_5_7_13.json'),
        os.path.join(path, '1D', 'P14', '1D_N16383_P14_F3_43_127.json'),
        os.path.join(path, '1D', 'P16', '1D_N65535_P16_F3_5_17_257.json'),
        os.path.join(path, '1D', 'P18', '1D_N262143_P18_F3^3_7_19_73.json')
    ]

    period = 6

    wb = Workbook()
    wb.remove(wb.active)
    for file_path in file_paths:
        print(f'Period: {period:02d}')

        with open(file_path, 'rt') as fp:
            content = json.load(fp)

        ws = wb.create_sheet(f'P{period:02d}')

        ws.cell(row=1, column=1, value='pos')
        ws.cell(row=1, column=2, value='time')
        ws.cell(row=1, column=3, value='count')

        last = content['spaces'][str(period)]
        row = 2
        for item in last:
            pos = item['pos'][0]
            time = item['time']
            count = item['count']
            ws.cell(row=row, column=1, value=pos)
            ws.cell(row=row, column=2, value=time)
            ws.cell(row=row, column=3, value=count)
            row += 1
        
        period += 2

    wb.save(os.path.join(path, '1D', 'times_new.xlsx'))

def main2():
    upper = 10000
    divupper = upper / 2.0

    print('   n  p1d  p2d  p3d')
    print('---- ---- ---- ----')
    num_p1d = 0
    num_p2d = 0
    num_p3d = 0
    for n in range(1, upper+1, 2):
        p1d = getPeriod(n, 2)
        p2d = getPeriod(n, 4)
        p3d = getPeriod(n, 8)
        # print(f'{n:4d} {p1d:4d} {p2d:4d} {p3d:4d}')

        if p1d % 2 == 1:
            num_p1d += 1
        if p2d % 2 == 1:
            num_p2d += 1
        if p3d % 2 == 1:
            num_p3d += 1
        
    print('')
    print(f'Odd period 1d: {100.0 * num_p1d / divupper:5.2f}%')
    print(f'Odd period 2d: {100.0 * num_p2d / divupper:5.2f}%')
    print(f'Odd period 3d: {100.0 * num_p3d / divupper:5.2f}%')


if __name__ == '__main__':
    main2()
